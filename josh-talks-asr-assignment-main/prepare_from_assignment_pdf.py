import argparse
import json
import re
from pathlib import Path
from urllib.parse import urlparse

import openpyxl
import requests
from pypdf import PdfReader

from q2_cleanup_pipeline import build_q2_live_report
from q3_spell_checker import run_spell_check_pipeline, build_q3_live_report


def extract_urls_from_pdf(pdf_path: Path) -> list[str]:
    reader = PdfReader(str(pdf_path))
    urls: set[str] = set()
    pattern = re.compile(r"https?://[^\s)\]>\"']+")

    for page in reader.pages:
        text = page.extract_text() or ""
        urls.update(pattern.findall(text))

        annotations = page.get("/Annots") or []
        for annotation in annotations:
            obj = annotation.get_object()
            action = obj.get("/A")
            if action and action.get("/URI"):
                urls.add(str(action.get("/URI")))

    return sorted(urls)


def download_assignment_sheets(urls: list[str], out_dir: Path) -> list[Path]:
    out_dir.mkdir(parents=True, exist_ok=True)
    sheet_urls = [u for u in urls if "docs.google.com/spreadsheets" in u]
    downloaded: list[Path] = []

    for url in sheet_urls:
        match = re.search(r"/d/([^/]+)/", url)
        if not match:
            continue
        sheet_id = match.group(1)
        xlsx_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"
        response = requests.get(xlsx_url, timeout=30)
        response.raise_for_status()
        out_path = out_dir / f"{sheet_id}.xlsx"
        out_path.write_bytes(response.content)
        downloaded.append(out_path)

    return downloaded


def find_workbooks(workbooks: list[Path]) -> tuple[Path, Path, Path]:
    word_wb = None
    dataset_wb = None
    task_wb = None

    for workbook_path in workbooks:
        workbook = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            rows = list(sheet.iter_rows(min_row=1, max_row=2, values_only=True))
            if not rows:
                continue

            header = [str(c).strip() if c is not None else "" for c in rows[0]]
            header_set = set(header)

            if "word" in header_set:
                word_wb = workbook_path

            if {"user_id", "recording_id", "language", "duration", "rec_url_gcp"}.issubset(header_set):
                dataset_wb = workbook_path

            if "Human" in header_set and any(h.startswith("Model ") for h in header if h):
                task_wb = workbook_path

    if not word_wb or not dataset_wb or not task_wb:
        raise RuntimeError(
            f"Could not identify all required sheets. word={word_wb}, dataset={dataset_wb}, task={task_wb}"
        )

    return word_wb, dataset_wb, task_wb


def build_manifest(dataset_workbook: Path, output_manifest: Path) -> int:
    workbook = openpyxl.load_workbook(dataset_workbook, read_only=True, data_only=True)

    target_sheet = None
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        header = [str(c).strip() if c is not None else "" for c in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
        if {"user_id", "recording_id", "language", "duration", "rec_url_gcp"}.issubset(set(header)):
            target_sheet = sheet
            break

    if target_sheet is None:
        raise RuntimeError("Dataset sheet with manifest columns not found")

    rows = list(target_sheet.iter_rows(values_only=True))
    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    index = {name: i for i, name in enumerate(header) if name}

    manifest_records = []
    for row in rows[1:]:
        if not row or row[index["recording_id"]] is None:
            continue

        old_rec = str(row[index["rec_url_gcp"]] or "").strip()
        recording_id = str(int(row[index["recording_id"]]))
        language = row[index["language"]] or "hi"
        duration = float(row[index["duration"]] or 0)
        user_id = int(row[index["user_id"]])

        parts = urlparse(old_rec).path.strip("/").split("/")
        folder_id = parts[-2] if len(parts) >= 2 else None
        if not folder_id:
            continue

        manifest_records.append(
            {
                "user_id": user_id,
                "recording_id": int(recording_id),
                "language": language,
                "duration": duration,
                "rec_url_gcp": f"https://storage.googleapis.com/upload_goai/{folder_id}/{recording_id}_audio.wav",
                "transcription_url": f"https://storage.googleapis.com/upload_goai/{folder_id}/{recording_id}_transcription.json",
                "metadata_url": f"https://storage.googleapis.com/upload_goai/{folder_id}/{recording_id}_metadata.json",
            }
        )

    output_manifest.write_text(json.dumps(manifest_records, ensure_ascii=False, indent=2), encoding="utf-8")
    return len(manifest_records)


def build_q3_from_word_sheet(word_workbook: Path) -> dict:
    workbook = openpyxl.load_workbook(word_workbook, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))

    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    try:
        word_index = header.index("word")
    except ValueError as exc:
        raise RuntimeError("word column not found in word sheet") from exc

    words = []
    seen = set()
    for row in rows[1:]:
        if row is None:
            continue
        token = (str(row[word_index]).strip() if row[word_index] is not None else "")
        if token and token not in seen:
            seen.add(token)
            words.append(token)

    summary = run_spell_check_pipeline(words, output_csv="spell_check_results.csv")
    report = build_q3_live_report(results_csv=summary["output_file"], output_dir="./artifacts/q3", low_conf_sample_size=50)
    return report


def build_q2_q4_from_task_sheet(task_workbook: Path) -> tuple[dict, dict]:
    from q4_lattice_wer import LatticeBuilder, LatticeWERComputer

    workbook = openpyxl.load_workbook(task_workbook, read_only=True, data_only=True)

    target_sheet = None
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        header = [str(c).strip() if c is not None else "" for c in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
        if "Human" in set(header) and any(h.startswith("Model ") for h in header if h):
            target_sheet = sheet
            break

    if target_sheet is None:
        raise RuntimeError("Task sheet not found")

    rows = list(target_sheet.iter_rows(values_only=True))
    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    index = {name: i for i, name in enumerate(header) if name}

    human_col = index["Human"]
    model_cols = [h for h in header if h.startswith("Model ")]

    q2_pairs = []
    builder = LatticeBuilder()
    computer = LatticeWERComputer()
    std_sum = {m: 0.0 for m in model_cols}
    lat_sum = {m: 0.0 for m in model_cols}
    first_ref = None
    first_models = None
    count = 0

    for row in rows[1:]:
        reference = str(row[human_col]).strip() if row[human_col] is not None else ""
        if not reference:
            continue

        model_outputs = {}
        for model_name in model_cols:
            value = row[index[model_name]]
            if value is None:
                continue
            hypothesis = str(value).strip()
            if hypothesis:
                model_outputs[model_name] = hypothesis

        if not model_outputs:
            continue

        if "Model H" in model_outputs:
            q2_pairs.append((model_outputs["Model H"], reference))

        if len(model_outputs) < 2:
            continue

        lattice = builder.build(reference, model_outputs)
        for model_name, hypothesis in model_outputs.items():
            std = float(computer.compute_standard_wer(reference, hypothesis))
            lat = float(computer.compute(lattice, hypothesis)["wer"])
            std_sum[model_name] += std
            lat_sum[model_name] += lat

        count += 1
        if first_ref is None:
            first_ref = reference
            first_models = model_outputs

    q2_report = build_q2_live_report(q2_pairs, output_dir="./artifacts/q2", max_examples=5)

    q4_results = {}
    if count > 0:
        for model_name in model_cols:
            std = round(std_sum[model_name] / count, 2)
            lat = round(lat_sum[model_name] / count, 2)
            q4_results[model_name] = {
                "standard_wer": float(std),
                "lattice_wer": float(lat),
                "substitutions": None,
                "deletions": None,
                "insertions": None,
                "improved": bool(lat < std),
            }

    q4_report = {
        "source": "Assignment Task sheet (aggregate)",
        "n_utterances": int(count),
        "reference": first_ref or "",
        "models": first_models or {},
        "alignment_unit": "word",
        "trust_threshold": float(builder.TRUST_THRESHOLD),
        "results": q4_results,
    }

    q4_dir = Path("./artifacts/q4")
    q4_dir.mkdir(parents=True, exist_ok=True)
    (q4_dir / "report.json").write_text(json.dumps(q4_report, ensure_ascii=False, indent=2), encoding="utf-8")

    return q2_report, q4_report


def main() -> None:
    parser = argparse.ArgumentParser(description="Prepare assignment artifacts from the provided PDF links")
    parser.add_argument(
        "--pdf",
        default="New_Task Assignment _ AI Researcher Intern- Speech & Audio _ Josh Talks .pdf",
        help="Path to assignment PDF",
    )
    args = parser.parse_args()

    pdf_path = Path(args.pdf)
    if not pdf_path.exists():
        raise FileNotFoundError(f"PDF not found: {pdf_path}")

    urls = extract_urls_from_pdf(pdf_path)
    Path("assignment_urls.txt").write_text("\n".join(urls), encoding="utf-8")

    downloads_dir = Path("pdf_links_downloads")
    workbooks = download_assignment_sheets(urls, downloads_dir)

    word_wb, dataset_wb, task_wb = find_workbooks(workbooks)
    manifest_count = build_manifest(dataset_wb, Path("manifest.json"))
    q3_report = build_q3_from_word_sheet(word_wb)
    q2_report, q4_report = build_q2_q4_from_task_sheet(task_wb)

    print("\n=== ASSIGNMENT PREP SUMMARY ===")
    print(f"PDF: {pdf_path}")
    print(f"URLs extracted: {len(urls)}")
    print(f"Manifest records: {manifest_count}")
    print(f"Q2 report pairs: {q2_report.get('n_pairs')}")
    print(f"Q3 total words: {q3_report.get('total_words')}")
    print(f"Q4 utterances: {q4_report.get('n_utterances')}")
    print("Artifacts ready: artifacts/q2/report.json, artifacts/q3/report.json, artifacts/q4/report.json")
    print("Q1 remains training-dependent. Run q1_whisper_finetune.py to produce artifacts/q1/report.json")


if __name__ == "__main__":
    main()
